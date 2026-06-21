#!/usr/bin/env python3
"""
recalculate_baseline_gabungan.py
────────────────────────────────
Menghitung ulang 'Rata-rata Omzet' dan 'Rata-rata Order' dari file
BASELINE_GABUNGAN yang sudah ada dengan logika yang BENAR:

  ✅ Akumulasi semua portal per bulan terlebih dahulu,
     baru hasil akumulasi tersebut dirata-rata.

  ❌ BUKAN: rata-rata per portal terlebih dahulu, lalu dijumlah.

Hasil per-baris (per-portal) TIDAK diubah — hanya kolom 'Rata-rata'
di masing-masing baris yang diperbarui, dan ringkasan platform
dicetak ke layar (dan disimpan ke sheet 'Summary').

Usage:
  python recalculate_baseline_gabungan.py
  python recalculate_baseline_gabungan.py --path /path/ke/BASELINE_GABUNGAN.xlsx
"""

import argparse
import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
from datetime import datetime


DEFAULT_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "laporan", "baseline", "2026-02-01_to_2026-04-30",
    "BASELINE_GABUNGAN_SUPERFOOD.xlsx"
)


def recalculate(filepath: str, dry_run: bool = False):
    if not os.path.exists(filepath):
        print(f"❌ File tidak ditemukan: {filepath}")
        sys.exit(1)

    print(f"\n📂 Membaca file: {filepath}")
    df = pd.read_excel(filepath, sheet_name="Baseline Summary")

    if df.empty:
        print("❌ Sheet 'Baseline Summary' kosong.")
        sys.exit(1)

    print(f"✅ {len(df)} baris data ditemukan.")
    print(f"   Kolom: {list(df.columns)}\n")

    # ── Identifikasi kolom bulan ──────────────────────────────────────────
    order_month_cols = sorted([c for c in df.columns if str(c).startswith("Order Bulan ke-")])
    omzet_month_cols = sorted([c for c in df.columns if str(c).startswith("Omzet Bulan ke-")])

    if not order_month_cols:
        print("❌ Tidak ditemukan kolom 'Order Bulan ke-N'. "
              "Pastikan file GABUNGAN memiliki kolom per-bulan.")
        sys.exit(1)

    num_months = len(order_month_cols)
    print(f"📅 Jumlah bulan terdeteksi : {num_months}")
    print(f"   Order  cols: {order_month_cols}")
    print(f"   Omzet  cols: {omzet_month_cols}\n")

    # ── Konversi ke numerik ───────────────────────────────────────────────
    for col in order_month_cols + omzet_month_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # ── Mask platform ─────────────────────────────────────────────────────
    app_lower   = df["Aplikasi"].astype(str).str.lower().str.strip()
    grab_mask   = app_lower.str.contains("grab",   na=False)
    shopee_mask = app_lower.str.contains("shopee", na=False)
    go_mask     = (~grab_mask) & (~shopee_mask) & app_lower.str.contains("go", na=False)

    def _compute_platform_avg(mask, cols):
        """
        Akumulasi semua portal (baris) per bulan terlebih dahulu (axis=0),
        lalu rata-rata akumulasi tersebut dibagi num_months.
        """
        grp = df.loc[mask, cols].copy()
        if grp.empty or not cols:
            return 0.0, {}
        monthly_totals = grp.sum(axis=0)           # jumlah semua portal per bulan
        detail = {col: monthly_totals[col] for col in cols}
        return float(monthly_totals.sum()) / num_months, detail

    # ── Hitung rata-rata BENAR per platform ───────────────────────────────
    avg_order_gr, detail_order_gr = _compute_platform_avg(grab_mask,   order_month_cols)
    avg_order_sf, detail_order_sf = _compute_platform_avg(shopee_mask, order_month_cols)
    avg_order_go, detail_order_go = _compute_platform_avg(go_mask,     order_month_cols)

    avg_omzet_gr, detail_omzet_gr = _compute_platform_avg(grab_mask,   omzet_month_cols)
    avg_omzet_sf, detail_omzet_sf = _compute_platform_avg(shopee_mask, omzet_month_cols)
    avg_omzet_go, detail_omzet_go = _compute_platform_avg(go_mask,     omzet_month_cols)

    # ── Cetak ringkasan ───────────────────────────────────────────────────
    def fmt_rp(v):
        return f"Rp {int(round(v)):,}".replace(",", ".")

    print("=" * 55)
    print("  📊 HASIL RECALCULATE (LOGIKA BENAR)")
    print("=" * 55)

    platforms = [
        ("GoFood",     go_mask,     avg_omzet_go, avg_order_go,
         detail_omzet_go, detail_order_go),
        ("GrabFood",   grab_mask,   avg_omzet_gr, avg_order_gr,
         detail_omzet_gr, detail_order_gr),
        ("ShopeeFood", shopee_mask, avg_omzet_sf, avg_order_sf,
         detail_omzet_sf, detail_order_sf),
    ]

    summary_rows = []
    for name, mask, avg_omzet, avg_order, d_omzet, d_order in platforms:
        n_portals = int(mask.sum())
        print(f"\n  🏪 {name}  ({n_portals} portal)")
        print(f"     Detail order per bulan :")
        for col, val in d_order.items():
            print(f"       {col}: {int(val):,}")
        total_order_all = sum(d_order.values())
        print(f"     Total order (semua bulan + portal): {int(total_order_all):,}")
        print(f"     ÷ {num_months} bulan  →  Rata-rata Order : {round(avg_order)}")
        print(f"     Rata-rata Omzet : {fmt_rp(avg_omzet)}")
        summary_rows.append({
            "Platform": name,
            "Jumlah Portal": n_portals,
            "Rata-rata Omzet": int(round(avg_omzet)),
            "Rata-rata Order": round(avg_order),
        })

    print("\n" + "=" * 55)
    print("  📋 RINGKASAN (copy untuk PDF / Discord)")
    print("=" * 55)
    print("\n📊 Rata-rata Omzet")
    for r in summary_rows:
        print(f"  {r['Platform']}: {fmt_rp(r['Rata-rata Omzet'])}")
    print("\n🛒 Rata-rata Order")
    for r in summary_rows:
        print(f"  {r['Platform']}: {r['Rata-rata Order']}")
    print()

    if dry_run:
        print("ℹ️  --dry-run aktif. File tidak dimodifikasi.")
        return

    # ── Update kolom 'Rata-rata Order' & 'Rata-rata Omzet' per baris ─────
    # Setiap baris tetap menyimpan rata-rata miliknya sendiri (per portal).
    # Yang berubah adalah CARA kita membaca agregat di cli.py (sudah difix).
    # Di sini kita cukup memastikan kolom tersebut benar per-portal.
    for col in order_month_cols + omzet_month_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    if "Rata-rata Omzet" in df.columns and "Rata-rata Order" in df.columns:
        for idx in df.index:
            omzet_cols_vals = [df.at[idx, c] for c in omzet_month_cols]
            order_cols_vals = [df.at[idx, c] for c in order_month_cols]
            df.at[idx, "Rata-rata Omzet"] = int(sum(omzet_cols_vals) / num_months)
            df.at[idx, "Rata-rata Order"] = round(sum(order_cols_vals) / num_months)

    # ── Tulis kembali ke file (sheet Baseline Summary diperbarui) ─────────
    summary_df = pd.DataFrame(summary_rows)

    print(f"💾 Menyimpan hasil ke: {filepath}")
    with pd.ExcelWriter(filepath, engine="openpyxl", mode="a",
                        if_sheet_exists="replace") as writer:
        df.to_excel(writer, index=False, sheet_name="Baseline Summary")
        summary_df.to_excel(writer, index=False, sheet_name="Platform Summary")

    print("✅ File berhasil diperbarui!")
    print(f"   • Sheet 'Baseline Summary' — data per-portal diperbarui")
    print(f"   • Sheet 'Platform Summary' — ringkasan agregat per platform\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Recalculate Baseline GABUNGAN averaging")
    parser.add_argument(
        "--path", type=str, default=DEFAULT_FILE,
        help="Path ke file BASELINE_GABUNGAN_*.xlsx"
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Hanya cetak hasil tanpa menyimpan perubahan"
    )
    args = parser.parse_args()
    recalculate(args.path, dry_run=args.dry_run)
